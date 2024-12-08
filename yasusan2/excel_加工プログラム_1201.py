# %%
import pandas as pd
from tqdm import tqdm
import numpy as np
from io import StringIO
import openpyxl as px
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
import warnings
import os
import glob
from openpyxl.styles import Border, Side

# %%
print('プログラムを開始します。')

# %%
waku_color_5 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow'}


waku_color_6 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green'}

waku_color_7 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange'}

waku_color_8 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange', '8':'pink'}

waku_color_9 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange', '8':'pink', '9':'pink'}

waku_color_10 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange',  '8':'orange', '9':'pink', '10':'pink'}

waku_color_11 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'green', '8':'orange',  '9':'orange', '10':'pink', '11':'pink'}

waku_color_12 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow', '6':'yellow',
                '7':'green', '8':'green', '9':'orange',  '10':'orange', '11':'pink', '12':'pink'}

waku_color_13 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'blue', '6':'yellow', '7':'yellow',
                '8':'green', '9':'green', '10':'orange',  '11':'orange', '12':'pink', '13':'pink'}

waku_color_14 = {'1':'white', '2':'black', '3':'red', '4':'red', '5':'blue', '6':'blue', '7':'yellow', '8':'yellow',
                '9':'green', '10':'green', '11':'orange',  '12':'orange', '13':'pink', '14':'pink'}

waku_color_15 = {'1':'white', '2':'black', '3':'black', '4':'red', '5':'red', '6':'blue', '7':'blue', '8':'yellow', '9':'yellow',
                '10':'green', '11':'green', '12':'orange',  '13':'orange', '14':'pink', '15':'pink'}

waku_color_16 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'pink', '16':'pink'}

waku_color_17 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'pink', '16':'pink', '17':'pink'}

waku_color_18 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'orange', '16':'pink', '17':'pink', '18':'pink'}

waku_color_dic = {5:waku_color_5, 6:waku_color_6, 7:waku_color_7, 8:waku_color_8, 9:waku_color_9, 10:waku_color_10,
                11:waku_color_11, 12:waku_color_12, 13:waku_color_13, 14:waku_color_14, 15:waku_color_15, 16:waku_color_16,
                17:waku_color_17, 18:waku_color_18}
colors = {
    "white": "FFFFFF",
    "black": "000000",
    "red": "FF0000",
    "blue": "0000FF",
    "yellow": "FFFF00",
    "green": "00FF00",
    "orange": "FFA500",
    "pink": "FFC0CB"
}
# Define the border style
thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

# %%
os.makedirs('加工済みファイル', exist_ok=True)

# %%
columns_list = [4, 5, 7, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
columns_name_list = ['馬齢', 'オッズ', '斤量', '脚質', '総合 値', 'SP 値', 'AG 値', 'SA 値', '馬連率', '戦数', '賞金平均', 'KI 値', '総合値']
#columns_name_list = ['馬齢', '斤量', '脚質', '総合 値', 'SP 値', 'AG 値', 'SA 値', '馬連率', '戦数', '賞金平均', 'KI 値', '総合値']

# %%
# A列の最終行を取得する
def get_last_row_in_column(sheet, column):
    last_row = 0
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=column).value is not None:
            last_row = row
    return last_row

# %%
def excel_prosessing(sheetname):
    ws = wb[sheetname]
    last_row = get_last_row_in_column(ws, 1)
    last_row_b = get_last_row_in_column(ws, 6)
    # シートの最終列を取得する
    last_column = ws.max_column
    #馬番だけは外だし
    list_umaban = []
    for row in range(2, last_row + 1):
        list_umaban.append(ws.cell(row, 2).value)
    list_umaban
    df_umaban = pd.DataFrame(list_umaban)
    # 1行目を列名に設定
    df_umaban.columns = df_umaban.iloc[0]
    # オッズの列を取得
    list_odds = []
    for row in range(2, last_row + 1):
        list_odds.append(ws.cell(row, 5).value)
    df_odds = pd.DataFrame(list_odds)
    # 1行目を列名に設定
    df_odds.columns = df_odds.iloc[0]

    # 1行目を削除
    df_umaban = df_umaban.drop(df_umaban.index[0])
    df_odds = df_odds.drop(df_odds.index[0])
    list_df = []
    only_columns_list = []
    for column in range(1,last_column+1):
        if  ws.cell(2, column).value in columns_name_list:
            if ws.cell(2, column).value != 'オッズ':
                list_data = []    
                for row in range(2, last_row + 1):
                    list_data.append(ws.cell(row, column).value)
                df = pd.DataFrame(list_data)
                # 1行目を列名に設定
                df.columns = df.iloc[0]
                df = df.drop(df.index[0])
                df_concat = pd.concat([df_umaban, df, df_odds], axis=1)
                df_concat = df_concat.sort_values(by=df_concat.columns[2], ascending=True).reset_index(drop=True)
                df_concat = df_concat[df_concat.columns[:-1]]
                list_df.append(df_concat)
                only_columns_list.append(ws.cell(2, column).value)
            else:
                # 1行目を列名に設定
                df_concat = pd.concat([df_umaban, df_odds], axis=1)
                df_concat = df_concat.sort_values(by=df_concat.columns[1], ascending=True).reset_index(drop=True)
                list_df.append(df_concat)
                only_columns_list.append(ws.cell(2, column).value)
    # 上の作成
    up_chart_list = []
    for a, df in enumerate(list_df):
        if a <= 3:
            up_chart_list.append(df.sort_values(by=df.columns[1], ascending=True).reset_index(drop=True))
        else:
            up_chart_list.append(df.sort_values(by=df.columns[1], ascending=False).reset_index(drop=True))
    #下の表作成
    down_chart_list = []
    for a, df in enumerate(list_df):
        if a <= 3:
            down_chart_list.append(df.sort_values(by=df.columns[1], ascending=False).reset_index(drop=True))
        else:
            down_chart_list.append(df.sort_values(by=df.columns[1], ascending=True).reset_index(drop=True))
    up_chart_df = pd.concat(up_chart_list, axis=1)
    down_chart_df = pd.concat(down_chart_list, axis=1)
    # シートに入力 上表
    for a, column in zip(range(2, 2 + len(up_chart_df.columns)), range(len(up_chart_df.columns))): #列を指定
        for b, data in zip(range(last_row_b+4, last_row_b+4 + len(up_chart_df)), range(len(up_chart_df.iloc[:, column])+1)): #行を指定
            ws.cell(b, a).value = up_chart_df.iloc[:, column][data]
            ws.cell(b, a).font = Font(bold=True, name='HGPゴシックE') 
    # シートに入力 下表
    for a, column in zip(range(2, 2 + len(up_chart_df.columns)), range(len(down_chart_df.columns))): #列を指定
        for b, data in zip(range(last_row_b+len(down_chart_df) + 7, last_row_b+len(down_chart_df) + len(down_chart_df) + 7), range(len(down_chart_df.iloc[:, column]))): #行を指定
            ws.cell(b, a).value = down_chart_df.iloc[:, column][data]
            ws.cell(b, a).font = Font(bold=True, name='HGPゴシックE')
    #列名をシートに入力
    for a, data in zip(range(2, 2 + len(up_chart_df.columns)), up_chart_df.columns):
        #上表
        ws.cell(last_row_b+3, a).value = data
        ws.cell(last_row_b+3, a).font = Font(bold=True, name='HGPゴシックE')
        ws.cell(last_row_b+3, a).border = thin_border
        #下表
        ws.cell(last_row_b+len(down_chart_df)+6, a).value = data
        ws.cell(last_row_b+len(down_chart_df)+6, a).font = Font(bold=True, name='HGPゴシックE')
        ws.cell(last_row_b+len(down_chart_df)+6, a).border = thin_border
        
    fill_color_dict = {}
    font_color_dict = {}
    for a in range(3, 3 + len(up_chart_df)):
        cell = ws['B'+str(a)]
        fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        font_color = cell.font.color.rgb if cell.font.color else None
        data = cell.value
        fill_color_dict[data] = fill_color
        font_color_dict[data] = font_color
    # 順位の色取得
    rank_color_dict = {}
    for a in range(3, 3 + len(up_chart_df)):
        rank = ws['C'+str(a)]
        umaban = ws['B'+str(a)]
        fill_color = rank.fill.fgColor.rgb if rank.fill.fgColor else None
        data = umaban.value
        if fill_color != '00000000':
            rank_color_dict[data] = fill_color
    #上の表の色付け
    for a in range(2, 2 + len(up_chart_df.columns)):
        data = ws.cell(last_row_b+3, a).value
        if data == '馬番':
            for b in range(last_row_b+4, last_row_b+4 + len(up_chart_df)):
                data = ws.cell(b, a).value
                ws.cell(b, a).fill = PatternFill(fill_type='solid', fgColor=fill_color_dict[data])
                ws.cell(b, a).font = Font(color=font_color_dict[data], bold=True, name='HGPゴシックE') #  bold=True 太文字
        else:
            for b in range(last_row_b+4, last_row_b+4 + len(up_chart_df)):
                data = ws.cell(b, a-1).value
                try:
                    ws.cell(b, a).fill = PatternFill(fill_type='solid', fgColor=rank_color_dict[data])
                    ws.cell(b, a).font = Font(bold=True, name='HGPゴシックE')
                except KeyError:
                    pass
    #下の表色付け
    for a in range(2, 2+len(down_chart_df.columns)):
        data = ws.cell(last_row_b+6 + len(up_chart_df), a).value
        if data == '馬番':
            for b in range(last_row_b+7 + len(up_chart_df), last_row_b+7 + len(up_chart_df)+ len(down_chart_df)):
                data = ws.cell(b, a).value
                ws.cell(b, a).fill = PatternFill(fill_type='solid', fgColor=fill_color_dict[data])
                ws.cell(b, a).font = Font(color=font_color_dict[data], bold=True, name='HGPゴシックE') #  bold=True 太文字
        else:
            for b in range(last_row_b+7 + len(up_chart_df), last_row_b+7 + len(up_chart_df)+ len(down_chart_df)):
                data = ws.cell(b, a-1).value
                try:
                    ws.cell(b, a).fill = PatternFill(fill_type='solid', fgColor=rank_color_dict[data])
                    ws.cell(b, a).font = Font(bold=True, name='HGPゴシックE')
                except KeyError:
                    pass
    for column in ['S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA']:
        adjusted_width = 5
        ws.column_dimensions[column].width = adjusted_width
    # 賞金平均の時には列の幅を変更する
    for a in range(1, last_column + 2 + len(up_chart_df.columns)):
        if ws.cell(2, a).value == '賞金平均':
            col_letter = ws.cell(1, a).column_letter
            ws.column_dimensions[col_letter].width = 7
        if ws.cell(last_row_b+3, a).value == '賞金平均':
            col_letter = ws.cell(1, a).column_letter
            ws.column_dimensions[col_letter].width = 7
    
    # Apply the border to cells from S2 to AD2
    # 枠線作成
    for col in range(column_index_from_string('B'), column_index_from_string('J') + 1):
        #上表
        cell = ws.cell(row=last_row_b+3, column=col)
        cell.border = thin_border
        #下表
        cell = ws.cell(row=last_row_b+3+len(up_chart_df), column=col)
        cell.border = thin_border


# %%
files = glob.glob('[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]_??.xlsx')

# %%
for file in tqdm(files):
    wb = px.open(file)
    sheetnames = wb.sheetnames
    for sheetname in sheetnames:
        excel_prosessing(sheetname)
        file_name = file.split('.')[0]
        wb.save('加工済みファイル\\' + file_name + '_加工済み.xlsx')


