# %%
from bs4 import BeautifulSoup
import requests
import datetime
import re
import pandas as pd
from tqdm import tqdm
import numpy as np
from io import StringIO
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
import warnings

# FutureWarningを無視する
warnings.filterwarnings('ignore', category=FutureWarning)


# %%
today =datetime.datetime.today()
today_year = str(today.year)
today_month = str(today.month)
today_monthz = str(today.month).zfill(2)
today_day = str(today.day)
today_dayz = str(today.day).zfill(2)
today_today = today_year +"/"+ today_month +"/"+ today_day
today_todayz = today_year +"/"+ today_monthz +"/"+ today_dayz
today_today2 = today_year + today_monthz + today_dayz
jra = today_year +'/'+ today_year +"/"+ today_month +"/"+ today_monthz + today_dayz
today_tw = today_month + '/' + today_day
day_of_week_str = today.strftime('%A')
dic_weekday = {'Monday':'月', 'Tuesday':'火', 'Wednesday':'水', 'Thursday':'木', 'Friday':'金', 'Saturday':'土', 'Sunday':'日'}


# %%
"""
if today.month >= 8 and today.day >= 31:
    raise SystemExit("今日の日付が7/31以降です。強制終了します。")
"""
# %%
dic_kaisai = {'12':'札幌', '22':'函館', '32':'福島', '42':'新潟', '52':'東京', 
            '62':'中山', '72':'中京', '82':'京都', '92':'阪神', '02':'小倉'}

waku_color_5 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow'}

waku_color_6 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green'}

waku_color_7 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange'}

waku_color_8 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange', '8':'pink'}

waku_color_9 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange', '8':'pink', '9':'pink'}

waku_color_10 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange',  '8':'orange', '9':'pink', '10':'pink'}

waku_color_11 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'green', '8':'orange',  '9':'orange', '10':'pink', '11':'pink'}

waku_color_12 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow', '6':'yellow',
                '7':'green', '8':'green', '9':'orange',  '10':'orange', '11':'pink', '12':'pink'}

waku_color_13 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'blue', '6':'yellow', '7':'yellow',
                '8':'green', '9':'green', '10':'orange',  '11':'orange', '12':'pink', '13':'pink'}

waku_color_14 = {'1':'white', '2':'black', '3':'red', '4':'red', '5':'blue', '6':'blue', '7':'yellow', '8':'yellow',
                '9':'green', '10':'green', '11':'orange',  '12':'orange', '13':'pink', '14':'pink'}

waku_color_15 = {'1':'white', '2':'black', '3':'black', '4':'red', '5':'red', '6':'blue', '7':'blue', '8':'yellow', '9':'yellow',
                '10':'green', '11':'green', '12':'orange',  '13':'orange', '14':'pink', '15':'pink'}

waku_color_16 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'pink', '16':'pink'}

waku_color_17 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'pink', '16':'pink', '17':'pink'}

waku_color_18 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'orange', '16':'pink', '17':'pink', '18':'pink'}

waku_color_dic = {5:waku_color_5, 6:waku_color_6, 7:waku_color_7, 8:waku_color_8, 9:waku_color_9, 10:waku_color_10,
                11:waku_color_11, 12:waku_color_12, 13:waku_color_13, 14:waku_color_14, 15:waku_color_15, 16:waku_color_16,
                17:waku_color_17, 18:waku_color_18}
colors = {
    "white": "FFFFFF",
    "black": "000000",
    "red": "FF0000",
    "blue": "0000FF",
    "yellow": "FFFF00",
    "green": "00FF00",
    "orange": "FFA500",
    "pink": "FFC0CB"
}


# %%
today_today2 = input('日付を入力してください(例:20210101)')

# %%
url = 'https://uma-x.jp/race'
res = requests.get(url)
res.encoding = res.apparent_encoding
soup = BeautifulSoup(res.text,'html.parser')

# %%
soup_a = soup.find_all('a', class_='top_race_menu')

# %%
race_data_list = []
kaijou_data_list = []
for race_url in soup_a:
    if today_today2 in race_url['href']:
        race_data_list.append(race_url['href'])
        kaijou = race_url['href'].split('/')[2][:2]
        if kaijou not in kaijou_data_list:
            kaijou_data_list.append(kaijou)

# %%
id_list = ['tbl_uma_no_1',
            'tbl_uma_no_2',
            'tbl_uma_no_3',
            'tbl_uma_no_4',
            'tbl_uma_no_5',
            'tbl_uma_no_6',
            'tbl_uma_no_7',
            'tbl_uma_no_8',
            'tbl_uma_no_9',
            'tbl_uma_no_10',
            'tbl_uma_no_11',
            'tbl_uma_no_12',
            'tbl_uma_no_13',
            'tbl_uma_no_14',
            'tbl_uma_no_15',
            'tbl_uma_no_16',
            'tbl_uma_no_17',
            'tbl_uma_no_18']

# %%
print('出馬表のスクレイピングを開始します')

# %%
shutuba_data_list = []
url_list = []
kekka_df_list = []
race_name_list = []
df_list = []
for kaijou in tqdm(kaijou_data_list):
    for race_data_url in race_data_list:
        try:
            if kaijou == race_data_url.split('/')[2][:2]:
                # 出馬表スクレイピング
                url = str('https://uma-x.jp'+str(race_data_url))
                url_2 = str('https://uma-x.jp'+str(race_data_url))
                res = requests.get(url)
                res.encoding = res.apparent_encoding
                soup = BeautifulSoup(res.text,'html.parser')
                df = pd.read_html(url)[1]
                sa_df = pd.read_html(url_2)[1]
                umaban_df = pd.read_html(url)[0]
                kyakushitu_list = []
                shoukyou_list = []
                for umaban_data in range(len(df)):
                    umaban = soup.find('tr', id=id_list[umaban_data])
                    try:
                        png_list = []
                        kyakushitu_ture = False
                        for index, a in enumerate(str(umaban.find('td', class_='h13')).split('<br/>')[0].split('>')[1:-2]):
                            #if 'kyakusitu_on.png' in a or 'kyakusitu_m.png' in a:
                            if 'kyakusitu_on.png' in a and kyakushitu_ture == False:
                                if index == 0:
                                    png_list.append(1)
                                    kyakushitu_ture = True
                                elif index == 1:
                                    png_list.append(2)
                                    kyakushitu_ture = True
                                elif index == 2:
                                    png_list.append(3)
                                    kyakushitu_ture = True
                                elif index == 3:
                                    png_list.append(4)
                                    kyakushitu_ture = True
                            elif 'kyakusitu_m.png' in a and kyakushitu_ture == False:
                                if index == 0:
                                    png_list.append(1)
                                    kyakushitu_ture = True
                                elif index == 1:
                                    png_list.append(2)
                                    kyakushitu_ture = True
                                elif index == 2:
                                    png_list.append(3)
                                    kyakushitu_ture = True
                                elif index == 3:
                                    png_list.append(4)
                                    kyakushitu_ture = True
                        if len(png_list) == 0:
                            kyakushitu_list.append('---')
                        elif len(png_list) == 1:
                            kyakushitu_list.append(png_list[0])
                        elif len(png_list) == 2:
                            kyakushitu_list.append(png_list[0] + 'or' + png_list[1])
                        elif len(png_list) == 3:
                            kyakushitu_list.append(png_list[0] + 'or' + png_list[1] + 'or' + png_list[2])
                        elif len(png_list) == 4:
                            kyakushitu_list.append(png_list[0] + 'or' + png_list[1] + 'or' + png_list[2] + 'or' + png_list[3])
                    except:
                        kyakushitu_list.append('---')
                df['脚質'] = kyakushitu_list
                #df['昇降 級'] = shoukyou_list
                umamei_list = []
                umarei_list = []
                for data in df['馬名']:
                    umamei_list.append(data.split(' ')[0]) # 馬名
                    umarei_list.append(data.split(' ')[2]) # 馬齢
                df['馬名'] = umamei_list
                df['馬齢'] = umarei_list
                df_concat = pd.concat([umaban_df, df], axis=1)
                try:
                    df_concat = df_concat[['馬 番', '馬名', '馬齢', '複勝単勝 オッズ', '騎手名', '斤量', '前回騎乗', '調教師', '脚質', '総合 値',
                            'SP 値', 'AG 値', 'SA 値', '馬 連率', '戦 数', '賞金 平均','KI 値']] #通常
                    df_concat['SA 値'] = sa_df['SA 値']
                    for i in df_concat.index:
                        df_concat.loc[i, '騎手名'] = re.sub(r'\d+', '', df_concat.loc[i, '騎手名'])
                        try:
                            df_concat.loc[i, '前回騎乗'] = re.sub(r'\d+', '', df_concat.loc[i, '前回騎乗'])
                        except:
                            df_concat.loc[i, '前回騎乗'] = np.nan
                        try:
                            df_concat.loc[i, '調教師'] = re.sub(r'\d+', '', df_concat.loc[i, '調教師'])
                        except:
                            df_concat.loc[i, '調教師'] = np.nan
                    df_concat.loc[:, '騎手名'] = df_concat['騎手名'].str.replace(' ', '')
                    df_concat.loc[:, '前回騎乗'] = df_concat['前回騎乗'].str.replace(' ', '')
                    df_concat.loc[:, '調教師'] = df_concat['調教師'].str.replace(' ', '')
                    df_concat.loc[:, '馬 連率'] = df_concat['馬 連率'].str.replace('%', '')
                    df_concat.loc[:, '馬 連率']  = pd.to_numeric(df_concat['馬 連率'],errors='coerce')
                    try:
                        df_concat.loc[:, '脚質'] = df_concat['脚質'].astype(str)
                        df_concat.loc[:, '脚質'] = df_concat['脚質'].str.replace('1', '1 逃').str.replace('2', '2 先').str.replace('3', '3 差').str.replace('4', '4 追')
                    except AttributeError:
                        pass
                    df_concat = df_concat.rename({'馬 番':'馬番','馬 連率':'馬連率','戦 数':'戦数','総合 値':'総合値', '賞金 平均':'賞金平均'}, axis=1)
                except:
                    try:
                        df_concat = df_concat[['馬 番', '馬名', '馬齢', '複勝単勝 オッズ', '騎手名', '斤量', '前回騎乗', '調教師', '脚質', '総合 値',
                            'SP 値', 'AG 値', '馬 連率', '戦 数', '賞金 平均','KI 値']] #障害
                        for i in df_concat.index:
                            df_concat.loc[i, '騎手名'] = re.sub(r'\d+', '', df_concat.loc[i, '騎手名'])
                            try:
                                df_concat.loc[i, '調教師'] = re.sub(r'\d+', '', df_concat.loc[i, '調教師'])
                            except:
                                df_concat.loc[i, '調教師'] = np.nan
                        try:
                            df_concat.loc[i, '前回騎乗'] = re.sub(r'\d+', '', df_concat.loc[i, '前回騎乗'])
                        except:
                            df_concat.loc[i, '前回騎乗'] = np.nan
                        df_concat.loc[:, '騎手名'] = df_concat['騎手名'].str.replace(' ', '')
                        df_concat.loc[:, '前回騎乗'] = df_concat['前回騎乗'].str.replace(' ', '')
                        df_concat.loc[:, '調教師'] = df_concat['調教師'].str.replace(' ', '')
                        df_concat.loc[:, '馬 連率'] = df_concat['馬 連率'].str.replace('%', '')
                        df_concat.loc[:, '馬 連率']  = pd.to_numeric(df_concat['馬 連率'],errors='coerce')
                        df_concat = df_concat.rename({'馬 番':'馬番','馬 連率':'馬連率','戦 数':'戦数', '賞金 平均':'賞金平均'}, axis=1)
                        df_concat.loc[:, '馬番']  = pd.to_numeric(df_concat['馬番'],errors='coerce')
                        try:
                            df_concat.loc[:, '脚質'] = df_concat['脚質'].astype(str)
                            df_concat.loc[:, '脚質'] = df_concat['脚質'].str.replace('1', '1 逃').str.replace('2', '2 先').str.replace('3', '3 差').str.replace('4', '4 追')
                        except AttributeError:
                            pass
                    except:
                        df_concat = df_concat[['馬 番', '馬名', '馬齢', '複勝単勝 オッズ', '騎手名', '斤量', '調教師', '脚質', '総合 値',]] # 新馬
                        for i in df_concat.index:
                            df_concat.loc[i, '騎手名'] = re.sub(r'\d+', '', df_concat.loc[i, '騎手名'])
                            try:
                                df_concat.loc[i, '調教師'] = re.sub(r'\d+', '', df_concat.loc[i, '調教師'])
                            except:
                                df_concat.loc[i, '調教師'] = np.nan
                        df_concat.loc[:, '騎手名'] = df_concat['騎手名'].str.replace(' ', '')
                        df_concat.loc[:, '調教師'] = df_concat['調教師'].str.replace(' ', '')
                        df_concat.loc[:, '脚質'] = df_concat['脚質'].astype(str)
                        df_concat.loc[:, '脚質'] = df_concat['脚質'].str.replace('1', '1 逃').str.replace('2', '2 先').str.replace('3', '3 差').str.replace('4', '4 追')
                        try:
                            df_concat = df_concat.rename({'馬 番':'馬番','戦 数':'戦数'}, axis=1)
                        except:
                            pass
                sp_list = []
                if 'SP 値' in df_concat.columns:
                    for a in df_concat['SP 値']:
                        if a == '-':
                            sp_list.append(0)
                        else:
                            try:
                                sp_list.append(int(str(a)[:-4]))
                            except ValueError:
                                sp_list.append(str(a)[:-4].replace('-',''))
                            except TypeError:
                                sp_list.append(0)
                                
                    df_concat['SP 値'] = sp_list
                    df_concat['SP 値'] = pd.to_numeric(df_concat['SP 値'], errors='coerce')
                    ag_list = []
                    for a in df_concat['AG 値'].astype(str):
                        if a == '-':
                            ag_list.append(0)
                        else:
                            try:
                                ag_list.append(int(a.split('.')[0][:-2]))
                            except ValueError:
                                ag_list.append(0)
                    df_concat['AG 値'] = ag_list
                    try:
                        sa_list = []
                        for a in df_concat['SA 値'].astype(str):
                            if a == '-':
                                sa_list.append(0)
                            else:
                                try:
                                    sa_list.append(int(a[:-2]))
                                except ValueError:
                                    sa_list.append(0)
                        df_concat['SA 値'] = sa_list
                    except:
                        pass
                odds_all_list = []
                for a in df_concat['複勝単勝 オッズ']:
                    try:
                        odds_list = a.split('.')
                        odds_all_list.append(odds_list[0]+'.'+odds_list[1][:1])
                    except:
                        odds_all_list.append(0)
                df_concat['複勝単勝 オッズ'] = odds_all_list
                df_concat['複勝単勝 オッズ'] = df_concat['複勝単勝 オッズ'].astype(float)
                df_sort = df_concat.sort_values(by='複勝単勝 オッズ', ascending=True)
                df_sort = df_sort.rename(columns={'複勝単勝 オッズ':'単勝オッズ'})
                df_sort = df_sort[df_sort['単勝オッズ'] != 0]
                    
                df_list.append(df_sort)
                url_list.append(url)
        except :
            pass

# %%
book = openpyxl.Workbook()

# %%
#エクセルに保存
print('エクセルに保存します')
book = openpyxl.Workbook()
for kaijou in tqdm(kaijou_data_list):
    with pd.ExcelWriter('出馬表_'+today_today2+'_'+dic_kaisai[kaijou]+'.xlsx') as writer:
        for url, data in zip(url_list,df_list):
            if kaijou == str(url.split('/')[4][:2]):
                shetname_list = '出馬表'+'_'+ str(int(url.split('/')[4][9:11]))
                data.to_excel(writer, sheet_name=shetname_list, index=False)

# %%
print('エクセルの色付け、フォント変更を行います')
for kaijou in tqdm(kaijou_data_list):
    url_list_2 = []
    for url in url_list:
        if kaijou == str(url.split('/')[4][:2]):
            url_list_2.append(url)
    wb = openpyxl.load_workbook('出馬表_'+today_today2+'_'+dic_kaisai[kaijou]+'.xlsx')
    for url, sheetname in zip(url_list_2, wb.sheetnames):
        res = requests.get(url)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text,'html.parser')
        race_name = soup.find('h3',class_='race_name').text
        race_type = str(soup.find('h4').text.split(' ')[1]) + str(soup.find('h4').text.split(' ')[2])
        ws = wb[sheetname]
        last_column = ws.max_column
        ws.insert_rows(0)
        ws.auto_filter.ref = 'A2:R5'
        #シートの最終行を取得
        maxRow = ws.max_row
        #行ループ
        #配列宣言
        My_Value = []
        for i in range(2,maxRow):
            #A列を配列へ格納
            try:
                My_Value.append(int(ws['A'+str(i)].value))
            except:
                pass
        ws['A1'] = dic_kaisai[kaijou]
        ws['B1'] = str(int(url.split('/')[4][9:11]))+'R '+ str(race_name)
        ws['C1'] = race_type
        ws['K1'] = url
        font_2 = Font(bold=True, name='HGPゴシックE') 
        for row in ws.iter_rows():
            for cell in row:
                ws[cell.coordinate].font = font_2
                row[5].number_format = '0.0' # 小数点1桁まで表示 
        for i in range(3, 21):
            data = str(ws['A'+str(i)].value)
            font = Font(color='FFFFFF', bold=True, name='HGPゴシックE') 
            if data == 'None':
                break
            else:
                try:
                    color = waku_color_dic[int(max(My_Value))][data]
                    if color == 'black' or color == 'red' or color == 'blue' or color == 'green':
                        fill = PatternFill(fill_type='solid', fgColor=colors[color])
                        ws['A'+str(i)].fill = fill
                        ws['A'+str(i)].font = font
                except:
                    pass
                else:
                    fill = PatternFill(fill_type='solid', fgColor=colors[color], bgColor='000000')
                    ws['A'+str(i)].fill = fill
        # 列幅調整
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            if column == 'A':
                adjusted_width = 5
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'B':
                adjusted_width = 20
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'C':
                adjusted_width = 8
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'D':
                adjusted_width = 7
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'E':
                adjusted_width = 14
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'F':
                adjusted_width = 10
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'G':
                adjusted_width = 14
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'H':
                adjusted_width = 11
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'I':
                adjusted_width = 5
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'P':
                adjusted_width = 6
                ws.column_dimensions[column].width = adjusted_width
            else:
                adjusted_width = 5
                ws.column_dimensions[column].width = adjusted_width
        wb.save('出馬表_'+today_today2+'_'+dic_kaisai[kaijou]+'.xlsx')
        for a in range(1, last_column + 2):
            if ws.cell(2, a).value == '賞金平均':
                col_letter = ws.cell(1, a).column_letter
                ws.column_dimensions[col_letter].width = 7
print('完了')

# %%
