# %%
print('プログラム開始')

# %%
from bs4 import BeautifulSoup
import requests
import datetime
import re
import pandas as pd
from tqdm import tqdm
import numpy as np
from io import StringIO
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
import warnings
import os
import glob
# FutureWarningを無視する
warnings.filterwarnings('ignore', category=FutureWarning)
kekka_files = ['結果_1.xlsx',
                '結果_2.xlsx',
                '結果_3.xlsx',
                '結果_4.xlsx',
                '結果_5.xlsx',
                '結果_6.xlsx',
                '結果_7.xlsx',
                '結果_8.xlsx',
                '結果_9.xlsx',
                '結果_10.xlsx',
                '結果_11.xlsx',
                '結果_12.xlsx']

# %%
"""
today =datetime.datetime.today()
if today.month >= 8 and today.day >=23:
    raise SystemExit("今日の日付が8/16以降です。強制終了します。")
"""
# %%
year = input('年月を入力してください(例:2021)')
dic_weekday = {'Monday':'月', 'Tuesday':'火', 'Wednesday':'水', 'Thursday':'木', 'Friday':'金', 'Saturday':'土', 'Sunday':'日'}

# %%
dic_kaisai = {'01':'12', '02':'22', '03':'32', '04':'42', '05':'52', '06':'62', '07':'72', '08':'82', '09':'92', '10':'02'}
dic_kaisai = {'12':'01', '22':'02', '32':'03', '42':'04', '52':'05', '62':'06', '72':'07', '82':'08', '92':'09', '02':'10'}

# %%
dic_kaisai = {'12':'札幌', '22':'函館', '32':'福島', '42':'新潟', '52':'東京', 
            '62':'中山', '72':'中京', '82':'京都', '92':'阪神', '02':'小倉'}

# %%
waku_color_5 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow'}

waku_color_6 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green'}

waku_color_7 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange'}

waku_color_8 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange', '8':'pink'}

waku_color_9 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange', '8':'pink', '9':'pink'}

waku_color_10 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'orange',  '8':'orange', '9':'pink', '10':'pink'}

waku_color_11 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow',
                '6':'green', '7':'green', '8':'orange',  '9':'orange', '10':'pink', '11':'pink'}

waku_color_12 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'yellow', '6':'yellow',
                '7':'green', '8':'green', '9':'orange',  '10':'orange', '11':'pink', '12':'pink'}

waku_color_13 = {'1':'white', '2':'black', '3':'red', '4':'blue', '5':'blue', '6':'yellow', '7':'yellow',
                '8':'green', '9':'green', '10':'orange',  '11':'orange', '12':'pink', '13':'pink'}

waku_color_14 = {'1':'white', '2':'black', '3':'red', '4':'red', '5':'blue', '6':'blue', '7':'yellow', '8':'yellow',
                '9':'green', '10':'green', '11':'orange',  '12':'orange', '13':'pink', '14':'pink'}

waku_color_15 = {'1':'white', '2':'black', '3':'black', '4':'red', '5':'red', '6':'blue', '7':'blue', '8':'yellow', '9':'yellow',
                '10':'green', '11':'green', '12':'orange',  '13':'orange', '14':'pink', '15':'pink'}

waku_color_16 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'pink', '16':'pink'}

waku_color_17 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'pink', '16':'pink', '17':'pink'}

waku_color_18 = {'1':'white', '2':'white', '3':'black', '4':'black', '5':'red', '6':'red', '7':'blue', '8':'blue', '9':'yellow', '10':'yellow',
                '11':'green', '12':'green', '13':'orange',  '14':'orange', '15':'orange', '16':'pink', '17':'pink', '18':'pink'}

waku_color_dic = {5:waku_color_5, 6:waku_color_6, 7:waku_color_7, 8:waku_color_8, 9:waku_color_9, 10:waku_color_10,
                11:waku_color_11, 12:waku_color_12, 13:waku_color_13, 14:waku_color_14, 15:waku_color_15, 16:waku_color_16,
                17:waku_color_17, 18:waku_color_18}
colors = {
    "white": "FFFFFF",
    "black": "000000",
    "red": "FF0000",
    "blue": "0000FF",
    "yellow": "FFFF00",
    "green": "00FF00",
    "orange": "FFA500",
    "pink": "FFC0CB"
}

# %% [markdown]
# # ウマークスのURLを取得

# %%
def excel_color(date, kaijou, url_list_2):
    # 最終的にデータを集結させるワークブック
    for file in kekka_files:
        try:
            wb = openpyxl.open(file)
            file_name = file
            break
        except:
            pass
    # ワークブックを作ったときに最初からあるワークシートを削除
    for file in kekka_files:
        if file == file_name:
            pass
        else:
            try:
                wb_2 = openpyxl.open(file)
                wb._sheets.append(wb_2._sheets[0])
                os.remove(file)
            except:
                pass
    os.remove(file_name)
    #wb.save(str(date)+'_'+dic_kaisai[kaijou]+'.xlsx')
    wb.save('concat_data.xlsx')
    wb = openpyxl.load_workbook('concat_data.xlsx')
    for url, sheetname in zip(url_list_2, wb.sheetnames):
        res = requests.get(url)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text,'html.parser')
        race_name = soup.find('h1',class_='race_name').text
        race_type = str(soup.find('h4').text.split(' ')[1]) + str(soup.find('h4').text.split(' ')[2])
        ws = wb[sheetname]
        ws.insert_rows(0)
        ws.auto_filter.ref = 'A2:R5'
        #シートの最終行を取得
        maxRow = ws.max_row
        #行ループ
        #配列宣言
        My_Value = []
        for i in range(2,maxRow):
            #B列を配列へ格納
            try:
                My_Value.append(int(ws.cell(i,2).value))
            except:
                pass
        ws['A1'] = dic_kaisai[kaijou]
        ws['B1'] = str(int(url.split('/')[4][9:11]))+'R'
        ws['C1'] = race_name
        ws['K1'] = url
        ws['D1'] = race_type
        #色付け
        font_2 = Font(bold=True, name='HGPゴシックE') 
        for row in ws.iter_rows():
            for cell in row:
                ws[cell.coordinate].font = font_2
                row[5].number_format = '0.0' # 小数点1桁まで表示 
        for i in range(3, 21):
            data = str(ws['B'+str(i)].value)
            font = Font(color='FFFFFF', bold=True, name='HGPゴシックE') 
            if data == 'None':
                break
            else:
                try:
                    color = waku_color_dic[int(max(My_Value))][data]
                    if color == 'black' or color == 'red' or color == 'blue' or color == 'green':
                        fill = PatternFill(fill_type='solid', fgColor=colors[color])
                        ws['B'+str(i)].fill = fill
                        ws['B'+str(i)].font = font
                except:
                    pass
                else:
                    fill = PatternFill(fill_type='solid', fgColor=colors[color], bgColor='000000')
                    ws['B'+str(i)].fill = fill
        # 列幅調整
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            if column == 'A':
                adjusted_width = 5
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'B':
                adjusted_width = 5
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'C':
                adjusted_width = 20
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'D':
                adjusted_width = 8
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'E':
                adjusted_width = 7
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'F':
                adjusted_width = 14
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'G':
                adjusted_width = 10
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'H':
                adjusted_width = 14
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'I':
                adjusted_width = 11
                ws.column_dimensions[column].width = adjusted_width
            elif column == 'Q':
                adjusted_width = 6
                ws.column_dimensions[column].width = adjusted_width
            else:
                adjusted_width = 5
                ws.column_dimensions[column].width = adjusted_width
        # 着順の色付け
        for i in range(3, ws.max_row):
            data = str(ws['A'+str(i)].value)
            if data == '1':
                fill = PatternFill(fill_type='solid', fgColor='FFC0CB')
                ws['C'+str(i)].fill = fill
                number = str(ws['B'+str(i)].value)
                for q in range(3, ws.max_row):
                    if number == str(ws['C'+str(q)].value):
                        ws['C'+str(q)].fill = fill
            elif data == '2':
                fill = PatternFill(fill_type='solid', fgColor='ADD8E6')
                ws['C'+str(i)].fill = fill
                number = str(ws['B'+str(i)].value)
                for q in range(3, ws.max_row):
                    if number == str(ws['C'+str(q)].value):
                        ws['C'+str(q)].fill = fill
            elif data == '3':
                fill = PatternFill(fill_type='solid', fgColor='ADFF2F')
                ws['C'+str(i)].fill = fill
                number = str(ws['B'+str(i)].value)
                for q in range(3, ws.max_row):
                    if number == str(ws['C'+str(q)].value):
                        ws['C'+str(q)].fill = fill
        #wb.save(str(date)+'_'+dic_kaisai[kaijou]+'.xlsx')
    os.remove('concat_data.xlsx')
    wb.save(url[-8:]+'_'+dic_kaisai[kaijou]+'.xlsx')

# %%
def make_excel_file(year):
    url = 'https://uma-x.jp/race/'+year
    res = requests.get(url)
    res.encoding = res.apparent_encoding
    soup = BeautifulSoup(res.text,'html.parser')
    soup_a = soup.find_all('a', class_='top_race_menu')
    race_data_list = []
    kaijou_data_list = []
    date_list = []
    kaijou_date_list = []
    for race_url in soup_a:
        try:
            kaijou = race_url['href'].split('/')[2][:2]
            race_data_list.append(race_url['href'])
            date = race_url['href'].split('/')[2][-5:]
            if kaijou not in kaijou_data_list:
                kaijou_data_list.append(kaijou)
            if int(date) not in date_list:
                date_list.append(int(date))
            kaijou_date_list.append(kaijou+'_'+date)
        except:
            pass
    date_list.sort()
    id_list = ['tbl_uma_no_1',
                'tbl_uma_no_2',
                'tbl_uma_no_3',
                'tbl_uma_no_4',
                'tbl_uma_no_5',
                'tbl_uma_no_6',
                'tbl_uma_no_7',
                'tbl_uma_no_8',
                'tbl_uma_no_9',
                'tbl_uma_no_10',
                'tbl_uma_no_11',
                'tbl_uma_no_12',
                'tbl_uma_no_13',
                'tbl_uma_no_14',
                'tbl_uma_no_15',
                'tbl_uma_no_16',
                'tbl_uma_no_17',
                'tbl_uma_no_18']
    # 出馬表の取得
    print('出馬表のスクレイピングを開始します')
    shutuba_data_list = []
    url_list = []
    kekka_df_list = []
    race_name_list = []
    for race_data_url in tqdm(race_data_list):
        # 出馬表スクレイピング
        url = str('https://uma-x.jp'+str(race_data_url).replace('race_result','race_kekka'))
        url_2 = str('https://uma-x.jp'+str(race_data_url))
        res = requests.get(url)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text,'html.parser')
        df = pd.read_html(url)[1]
        try:
            sa_df = pd.read_html(url_2)[1]
        except UnicodeDecodeError:
            sa_list = []
            res = requests.get(url_2)
            res.encoding = res.apparent_encoding
            soup = BeautifulSoup(res.text,'html.parser')
            tabel = soup.find_all('table', class_='tbl race_data_tbl')[1]
            for a in range(6,6+(len(df)*9),9):
                data =tabel.find_all('div', class_='relative')[a].text
                sa_list.append(data)
            sa_df = pd.DataFrame(sa_list)
            sa_df.columns = ['SA 値']
        umaban_df = pd.read_html(url)[0]
        umamei_list = []
        umarei_list = []
        for data in df['馬名']:
            umamei_list.append(data.split(' ')[0]) # 馬名
            umarei_list.append(data.split(' ')[2]) # 馬齢
        df['馬名'] = umamei_list
        df['馬齢'] = umarei_list
        df_concat = pd.concat([umaban_df, df], axis=1)
        try:
            df_concat = df_concat[['順 位','馬 番', '馬名', '馬齢', 'オッズ', '騎手名', '斤量', '前回騎乗', '調教師', '脚質', '総合 値',
                    'SP 値', 'AG 値', 'SA 値', '馬 連率', '戦 数', '賞金 平均','KI 値']]
            df_concat['SA 値'] = sa_df['SA 値']
            for i in df_concat.index:
                df_concat.loc[i, '騎手名'] = re.sub(r'\d+', '', df_concat.loc[i, '騎手名'])
                try:
                    df_concat.loc[i, '前回騎乗'] = re.sub(r'\d+', '', df_concat.loc[i, '前回騎乗'])
                except:
                    df_concat.loc[i, '前回騎乗'] = np.nan
                try:
                    df_concat.loc[i, '調教師'] = re.sub(r'\d+', '', df_concat.loc[i, '調教師'])
                except:
                    df_concat.loc[i, '調教師'] = np.nan
            df_concat.loc[:, '騎手名'] = df_concat['騎手名'].str.replace(' ', '')
            df_concat.loc[:, '前回騎乗'] = df_concat['前回騎乗'].str.replace(' ', '')
            df_concat.loc[:, '調教師'] = df_concat['調教師'].str.replace(' ', '')
            df_concat.loc[:, '馬 連率'] = df_concat['馬 連率'].str.replace('%', '')
            df_concat.loc[:, '馬 連率']  = pd.to_numeric(df_concat['馬 連率'],errors='coerce')
            try:
                df_concat.loc[:, '脚質'] = df_concat['脚質'].str.replace('逃', '1 逃').str.replace('先', '2 先').str.replace('差', '3 差').str.replace('追', '4 追')
            except AttributeError:
                pass
            df_concat = df_concat.rename({'馬 番':'馬番','馬 連率':'馬連率','戦 数':'戦数','総合 値':'総合値', '賞金 平均':'賞金平均'}, axis=1)
        except:
            try:
                df_concat = df_concat[['順 位','馬 番', '馬名', '馬齢', 'オッズ', '騎手名', '斤量', '前回騎乗', '調教師', '脚質', '総合 値',
                        'SP 値', 'AG 値', '馬 連率', '戦 数', '賞金 平均','KI 値']]
                for i in df_concat.index:
                    df_concat.loc[i, '騎手名'] = re.sub(r'\d+', '', df_concat.loc[i, '騎手名'])
                    try:
                        df_concat.loc[i, '前回騎乗'] = re.sub(r'\d+', '', df_concat.loc[i, '前回騎乗'])
                    except:
                        df_concat.loc[i, '前回騎乗'] = np.nan
                    try:
                        df_concat.loc[i, '調教師'] = re.sub(r'\d+', '', df_concat.loc[i, '調教師'])
                    except:
                        df_concat.loc[i, '調教師'] = np.nan
                df_concat.loc[:, '騎手名'] = df_concat['騎手名'].str.replace(' ', '')
                df_concat.loc[:, '前回騎乗'] = df_concat['前回騎乗'].str.replace(' ', '')
                df_concat.loc[:, '調教師'] = df_concat['調教師'].str.replace(' ', '')
                df_concat.loc[:, '馬 連率'] = df_concat['馬 連率'].str.replace('%', '')
                df_concat.loc[:, '馬 連率']  = pd.to_numeric(df_concat['馬 連率'],errors='coerce')
                df_concat = df_concat.rename({'馬 番':'馬番','馬 連率':'馬連率','戦 数':'戦数', '賞金 平均':'賞金平均'}, axis=1)
                try:
                    df_concat.loc[:, '脚質'] = df_concat['脚質'].str.replace('逃', '1 逃').str.replace('先', '2 先').str.replace('差', '3 差').str.replace('追', '4 追')
                except AttributeError:
                    pass
            except:
                try:
                    df_concat = df_concat[['順 位','馬 番', '馬名', '馬齢', 'オッズ', '騎手名', '斤量', '調教師', '脚質', '総合 値',]]
                    for i in df_concat.index:
                        df_concat.loc[i, '騎手名'] = re.sub(r'\d+', '', df_concat.loc[i, '騎手名'])
                        try:
                            df_concat.loc[i, '調教師'] = re.sub(r'\d+', '', df_concat.loc[i, '調教師'])
                        except:
                            df_concat.loc[i, '調教師'] = np.nan
                    df_concat.loc[:, '騎手名'] = df_concat['騎手名'].str.replace(' ', '')
                    df_concat.loc[:, '調教師'] = df_concat['調教師'].str.replace(' ', '')
                    try:
                        df_concat.loc[:, '脚質'] = df_concat['脚質'].str.replace('逃', '1 逃').str.replace('先', '2 先').str.replace('差', '3 差').str.replace('追', '4 追')
                    except AttributeError:
                        pass
                    df_concat = df_concat.rename({'馬 番':'馬番','戦 数':'戦数'}, axis=1)
                except: #新潟の落雷で中止になったレースを除外する
                    df_juni = pd.DataFrame()
                    df_odds = pd.DataFrame()
                    df_juni['順 位'] = df_concat['馬 番']
                    df_odds['オッズ'] = [0]*len(df_concat)
                    
                    df_concat = df_concat[['馬 番', '馬名', '馬齢', '騎手名', '斤量', '調教師', '脚質', '総合 値',]]
                    df_concat = pd.concat([df_juni, df_odds, df_concat], axis=1)
                    
                    df_concat = df_concat[['順 位','馬 番', '馬名', '馬齢', 'オッズ', '騎手名', '斤量', '調教師', '脚質', '総合 値',]]
                    for i in df_concat.index:
                        df_concat.loc[i, '騎手名'] = re.sub(r'\d+', '', df_concat.loc[i, '騎手名'])
                        try:
                            df_concat.loc[i, '調教師'] = re.sub(r'\d+', '', df_concat.loc[i, '調教師'])
                        except:
                            df_concat.loc[i, '調教師'] = np.nan
                    df_concat.loc[:, '騎手名'] = df_concat['騎手名'].str.replace(' ', '')
                    df_concat.loc[:, '調教師'] = df_concat['調教師'].str.replace(' ', '')
                    try:
                        df_concat.loc[:, '脚質'] = df_concat['脚質'].str.replace('逃', '1 逃').str.replace('先', '2 先').str.replace('差', '3 差').str.replace('追', '4 追')
                    except AttributeError:
                        pass
                    df_concat = df_concat.rename({'馬 番':'馬番','戦 数':'戦数'}, axis=1)
                    
        odds_all_list = []
        try:
            for a in df_concat['オッズ']:
                odds_list = a.split('.')
                odds_all_list.append(odds_list[0]+'.'+odds_list[1][:1])
            df_concat['オッズ'] = odds_all_list
        except AttributeError:
            pass
        df_concat['馬番'] = df_concat['馬番'].astype(int)
        sp_list = []
        if 'SP 値' in df_concat.columns:
            for a in df_concat['SP 値']:
                if a == '-':
                    sp_list.append(0)
                elif a == '0.00':
                    sp_list.append(0)
                else:
                    try:
                        sp_list.append(int(a[:-4]))
                    except ValueError:
                        sp_list.append(int(a[:-4].replace('-','')))
                    except TypeError:
                        sp_list.append(0)
            df_concat['SP 値'] = sp_list
            ag_list = []
            for a in df_concat['AG 値'].astype(str):
                if a == '-':
                    ag_list.append(0)
                else:
                    try:
                        ag_list.append(int(a.split('.')[0][:-2]))
                    except ValueError:
                        ag_list.append(0)
            df_concat['AG 値'] = ag_list
            try:
                sa_list = []
                for a in df_concat['SA 値'].astype(str):
                    if a == '-':
                        sa_list.append(0)
                    else:
                        try:
                            sa_list.append(int(a[:-2]))
                        except ValueError:
                            sa_list.append(0)
                df_concat['SA 値'] = sa_list
            except:
                pass
        df_sort = df_concat.sort_values(by='馬番', ascending=True)
        df_sort['順 位'] = pd.to_numeric(df_sort['順 位'], errors='coerce')
        kekka_df = df_sort.rename(columns={'順 位':'着順'}).sort_values(by='着順', ascending=True)
        
        empty_row = pd.DataFrame([None]*df_sort.shape[1]).T
        empty_row.columns = df_sort.columns
        for a in range(2):
            df_sort = pd.concat([df_sort, empty_row], ignore_index=True)
        
        kekka_df['オッズ'] = kekka_df['オッズ'].astype(float)
        kekka_df = kekka_df[kekka_df['オッズ']!=0]
        empty_row = pd.DataFrame([None]*kekka_df.shape[1]).T
        empty_row.columns = kekka_df.columns
        for a in range(2):
            kekka_df = pd.concat([kekka_df,empty_row], ignore_index=True)
        
        # 結果スクレイピング
        #url = 'https://uma-x.jp/race_kekka/5202403070620240622'
        res = requests.get(url)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text,'html.parser')
        kenshu_list = []
        umaban_list = []
        haraimodoshi_list = []
        ninki_list = []
            #単勝データ取得
        try:
            tanshou_umaban = soup.find_all('div',class_='tan no')[0].text.split('\n')
            tanshou_haraimodoshi = soup.find_all('div',class_='tan yen')[0].text.split('\n')
            tanshou_ninki = soup.find_all('div',class_='tan ninki')[0].text.split('\n')
            if len(tanshou_umaban) == 3:
                umaban_list.append(tanshou_umaban[1])
                haraimodoshi_list.append(tanshou_haraimodoshi[1])
                ninki_list.append(tanshou_ninki[1])
                kenshu_list.append('単勝')
            elif len(tanshou_umaban) == 4:
                umaban_list.append(tanshou_umaban[1])
                umaban_list.append(tanshou_umaban[2])
                haraimodoshi_list.append(tanshou_haraimodoshi[1])
                haraimodoshi_list.append(tanshou_haraimodoshi[2])
                ninki_list.append(tanshou_ninki[1])
                ninki_list.append(tanshou_ninki[2])
                kenshu_list.append('単勝')
                kenshu_list.append('単勝')
            #複勝データ取得
            hukushou_umaban = soup.find_all('div',class_='huku no')[0].text.split('\n')
            hukushou_haraimodoshi = soup.find_all('div',class_='huku yen')[0].text.split('\n')
            hukushou_ninki = soup.find_all('div',class_='huku ninki')[0].text.split('\n')
            if len(hukushou_umaban) == 4:
                umaban_list.append(hukushou_umaban[1])
                umaban_list.append(hukushou_umaban[2])
                haraimodoshi_list.append(hukushou_haraimodoshi[1])
                haraimodoshi_list.append(hukushou_haraimodoshi[2])
                ninki_list.append(hukushou_ninki[1])
                ninki_list.append(hukushou_ninki[2])
                kenshu_list.append('複勝')
                kenshu_list.append('複勝')
            elif len(hukushou_umaban) == 5:
                umaban_list.append(hukushou_umaban[1])
                umaban_list.append(hukushou_umaban[2])
                umaban_list.append(hukushou_umaban[3])
                haraimodoshi_list.append(hukushou_haraimodoshi[1])
                haraimodoshi_list.append(hukushou_haraimodoshi[2])
                haraimodoshi_list.append(hukushou_haraimodoshi[3])
                ninki_list.append(hukushou_ninki[1])
                ninki_list.append(hukushou_ninki[2])
                ninki_list.append(hukushou_ninki[3])
                kenshu_list.append('複勝')
                kenshu_list.append('複勝')
                kenshu_list.append('複勝')
            elif len(hukushou_umaban) == 6:
                umaban_list.append(hukushou_umaban[1])
                umaban_list.append(hukushou_umaban[2])
                umaban_list.append(hukushou_umaban[3])
                umaban_list.append(hukushou_umaban[4])
                haraimodoshi_list.append(hukushou_haraimodoshi[1])
                haraimodoshi_list.append(hukushou_haraimodoshi[2])
                haraimodoshi_list.append(hukushou_haraimodoshi[3])
                haraimodoshi_list.append(hukushou_haraimodoshi[4])
                ninki_list.append(hukushou_ninki[1])
                ninki_list.append(hukushou_ninki[2])
                ninki_list.append(hukushou_ninki[3])
                ninki_list.append(hukushou_ninki[4])
                kenshu_list.append('複勝')
                kenshu_list.append('複勝')
                kenshu_list.append('複勝')
                kenshu_list.append('複勝')
            #枠連データ取得
            try:
                wakuren_umaban = soup.find_all('div',class_='wakuren no')[0].text.split('\n')
                wakuren_haraimodoshi = soup.find_all('div',class_='wakuren yen')[0].text.split('\n')
                wakuren_ninki = soup.find_all('div',class_='wakuren ninki')[0].text.split('\n')
                if len(wakuren_umaban) == 3:
                    umaban_list.append(wakuren_umaban[1])
                    haraimodoshi_list.append(wakuren_haraimodoshi[1])
                    ninki_list.append(wakuren_ninki[1])
                    kenshu_list.append('枠連')
                elif len(wakuren_umaban) == 4:
                    umaban_list.append(wakuren_umaban[1])
                    umaban_list.append(wakuren_umaban[2])
                    haraimodoshi_list.append(wakuren_haraimodoshi[1])
                    haraimodoshi_list.append(wakuren_haraimodoshi[2])
                    ninki_list.append(wakuren_ninki[1])
                    ninki_list.append(wakuren_ninki[2])
                    kenshu_list.append('枠連')
                    kenshu_list.append('枠連')
                elif len(wakuren_umaban) == 5:
                    umaban_list.append(wakuren_umaban[1])
                    umaban_list.append(wakuren_umaban[2])
                    umaban_list.append(wakuren_umaban[3])
                    haraimodoshi_list.append(hukushou_haraimodoshi[1])
                    haraimodoshi_list.append(hukushou_haraimodoshi[2])
                    haraimodoshi_list.append(hukushou_haraimodoshi[3])
                    ninki_list.append(wakuren_ninki[1])
                    ninki_list.append(wakuren_ninki[2])
                    ninki_list.append(wakuren_ninki[3])
                    kenshu_list.append('枠連')
                    kenshu_list.append('枠連')
                    kenshu_list.append('枠連')
            except:
                pass
            #ワイドデータ取得
            wide_umaban = soup.find_all('div',class_='wide no')[0].text.split('\n')
            wide_haraimodoshi = soup.find_all('div',class_='wide yen')[0].text.split('\n')
            wide_ninki = soup.find_all('div',class_='wide ninki')[0].text.split('\n')
            if len(wide_umaban) == 4:
                umaban_list.append(wide_umaban[1])
                umaban_list.append(wide_umaban[2])
                haraimodoshi_list.append(wide_haraimodoshi[1])
                haraimodoshi_list.append(wide_haraimodoshi[2])
                ninki_list.append(wide_ninki[1])
                ninki_list.append(wide_ninki[2])
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
            elif len(wide_umaban) == 5:
                umaban_list.append(wide_umaban[1])
                umaban_list.append(wide_umaban[2])
                umaban_list.append(wide_umaban[3])
                haraimodoshi_list.append(wide_haraimodoshi[1])
                haraimodoshi_list.append(wide_haraimodoshi[2])
                haraimodoshi_list.append(wide_haraimodoshi[3])
                ninki_list.append(wide_ninki[1])
                ninki_list.append(wide_ninki[2])
                ninki_list.append(wide_ninki[3])
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
            elif len(wide_umaban) == 6:
                umaban_list.append(wide_umaban[1])
                umaban_list.append(wide_umaban[2])
                umaban_list.append(wide_umaban[3])
                umaban_list.append(wide_umaban[4])
                haraimodoshi_list.append(wide_haraimodoshi[1])
                haraimodoshi_list.append(wide_haraimodoshi[2])
                haraimodoshi_list.append(wide_haraimodoshi[3])
                haraimodoshi_list.append(wide_haraimodoshi[4])
                ninki_list.append(wide_ninki[1])
                ninki_list.append(wide_ninki[2])
                ninki_list.append(wide_ninki[3])
                ninki_list.append(wide_ninki[4])
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
            elif len(wide_umaban) == 7:
                umaban_list.append(wide_umaban[1])
                umaban_list.append(wide_umaban[2])
                umaban_list.append(wide_umaban[3])
                umaban_list.append(wide_umaban[4])
                umaban_list.append(wide_umaban[5])
                haraimodoshi_list.append(wide_haraimodoshi[1])
                haraimodoshi_list.append(wide_haraimodoshi[2])
                haraimodoshi_list.append(wide_haraimodoshi[3])
                haraimodoshi_list.append(wide_haraimodoshi[4])
                haraimodoshi_list.append(wide_haraimodoshi[5])
                ninki_list.append(wide_ninki[1])
                ninki_list.append(wide_ninki[2])
                ninki_list.append(wide_ninki[3])
                ninki_list.append(wide_ninki[4])
                ninki_list.append(wide_ninki[5])
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
                kenshu_list.append('ワイド')
            #馬連データ取得
            umaren_umaban = soup.find_all('div',class_='umaren no')[0].text.split('\n')
            umaren_haraimodoshi = soup.find_all('div',class_='umaren yen')[0].text.split('\n')
            umaren_ninki = soup.find_all('div',class_='umaren ninki')[0].text.split('\n')
            if len(umaren_umaban) == 3:
                umaban_list.append(umaren_umaban[1])
                haraimodoshi_list.append(umaren_haraimodoshi[1])
                ninki_list.append(umaren_ninki[1])
                kenshu_list.append('馬連')
            elif len(umaren_umaban) == 4:
                umaban_list.append(umaren_umaban[1])
                umaban_list.append(umaren_umaban[2])
                haraimodoshi_list.append(umaren_haraimodoshi[1])
                haraimodoshi_list.append(umaren_haraimodoshi[2])
                ninki_list.append(umaren_ninki[1])
                ninki_list.append(umaren_ninki[2])
                kenshu_list.append('馬連')
                kenshu_list.append('馬連')
            elif len(umaren_umaban) == 5:
                umaban_list.append(umaren_umaban[1])
                umaban_list.append(umaren_umaban[2])
                umaban_list.append(umaren_umaban[3])
                haraimodoshi_list.append(umaren_haraimodoshi[1])
                haraimodoshi_list.append(umaren_haraimodoshi[2])
                haraimodoshi_list.append(umaren_haraimodoshi[3])
                ninki_list.append(umaren_ninki[1])
                ninki_list.append(umaren_ninki[2])
                ninki_list.append(umaren_ninki[3])
                kenshu_list.append('馬連')
                kenshu_list.append('馬連')
                kenshu_list.append('馬連')
            #馬単データ取得
            umatan_umaban = soup.find_all('div',class_='umatan no')[0].text.split('\n')
            umatan_haraimodoshi = soup.find_all('div',class_='umatan yen')[0].text.split('\n')
            umatan_ninki = soup.find_all('div',class_='umatan ninki')[0].text.split('\n')
            if len(umatan_umaban) == 3:
                umaban_list.append(umatan_umaban[1])
                haraimodoshi_list.append(umatan_haraimodoshi[1])
                ninki_list.append(umatan_ninki[1])
                kenshu_list.append('馬単')
            elif len(umatan_umaban) == 4:
                umaban_list.append(umatan_umaban[1])
                umaban_list.append(umatan_umaban[2])
                haraimodoshi_list.append(umatan_haraimodoshi[1])
                haraimodoshi_list.append(umatan_haraimodoshi[2])
                ninki_list.append(umatan_ninki[1])
                ninki_list.append(umatan_ninki[2])
                kenshu_list.append('馬単')
                kenshu_list.append('馬単')
            elif len(umatan_umaban) == 5:
                umaban_list.append(umatan_umaban[1])
                umaban_list.append(umatan_umaban[2])
                umaban_list.append(umatan_umaban[3])
                haraimodoshi_list.append(umatan_haraimodoshi[1])
                haraimodoshi_list.append(umatan_haraimodoshi[2])
                haraimodoshi_list.append(umatan_haraimodoshi[3])
                ninki_list.append(umatan_ninki[1])
                ninki_list.append(umatan_ninki[2])
                ninki_list.append(umatan_ninki[3])
                kenshu_list.append('馬単')
                kenshu_list.append('馬単')
                kenshu_list.append('馬単')
            #3連複データ取得
            renpuku_umaban = soup.find_all('div',class_='sanrenhuku no')[0].text.split('\n')
            renpuku_haraimodoshi = soup.find_all('div',class_='sanrenhuku yen')[0].text.split('\n')
            renpuku_ninki = soup.find_all('div',class_='sanrenhuku ninki')[0].text.split('\n')
            if len(renpuku_umaban) == 3:
                umaban_list.append(renpuku_umaban[1])
                haraimodoshi_list.append(renpuku_haraimodoshi[1])
                ninki_list.append(renpuku_ninki[1])
                kenshu_list.append('3連複')
            elif len(renpuku_umaban) == 4:
                umaban_list.append(renpuku_umaban[1])
                umaban_list.append(renpuku_umaban[2])
                haraimodoshi_list.append(renpuku_haraimodoshi[1])
                haraimodoshi_list.append(renpuku_haraimodoshi[2])
                ninki_list.append(renpuku_ninki[1])
                ninki_list.append(renpuku_ninki[2])
                kenshu_list.append('3連複')
                kenshu_list.append('3連複')
            elif len(renpuku_umaban) == 5:
                umaban_list.append(renpuku_umaban[1])
                umaban_list.append(renpuku_umaban[2])
                umaban_list.append(renpuku_umaban[3])
                haraimodoshi_list.append(renpuku_haraimodoshi[1])
                haraimodoshi_list.append(renpuku_haraimodoshi[2])
                haraimodoshi_list.append(renpuku_haraimodoshi[3])
                ninki_list.append(renpuku_ninki[1])
                ninki_list.append(renpuku_ninki[2])
                ninki_list.append(renpuku_ninki[3])
                kenshu_list.append('3連複')
                kenshu_list.append('3連複')
                kenshu_list.append('3連複')
            #3連複データ取得
            rentan_umaban = soup.find_all('div',class_='sanrentan no')[0].text.split('\n')
            rentan_haraimodoshi = soup.find_all('div',class_='sanrentan yen')[0].text.split('\n')
            rentan_ninki = soup.find_all('div',class_='sanrentan ninki')[0].text.split('\n')
            if len(rentan_umaban) == 3:
                umaban_list.append(rentan_umaban[1])
                haraimodoshi_list.append(rentan_haraimodoshi[1])
                ninki_list.append(rentan_ninki[1])
                kenshu_list.append('3連単')
            elif len(rentan_umaban) == 4:
                umaban_list.append(rentan_umaban[1])
                umaban_list.append(rentan_umaban[2])
                haraimodoshi_list.append(rentan_haraimodoshi[1])
                haraimodoshi_list.append(rentan_haraimodoshi[2])
                ninki_list.append(rentan_ninki[1])
                ninki_list.append(rentan_ninki[2])
                kenshu_list.append('3連単')
                kenshu_list.append('3連単')
            elif len(rentan_umaban) == 5:
                umaban_list.append(rentan_umaban[1])
                umaban_list.append(rentan_umaban[2])
                umaban_list.append(rentan_umaban[3])
                haraimodoshi_list.append(rentan_haraimodoshi[1])
                haraimodoshi_list.append(rentan_haraimodoshi[2])
                haraimodoshi_list.append(rentan_haraimodoshi[3])
                ninki_list.append(rentan_ninki[1])
                ninki_list.append(rentan_ninki[2])
                ninki_list.append(rentan_ninki[3])
                kenshu_list.append('3連単')
                kenshu_list.append('3連単')
                kenshu_list.append('3連単')
            df_return = pd.DataFrame({'券種':kenshu_list, '馬番':umaban_list, '払戻':haraimodoshi_list, '人気':ninki_list})
            df_return['人気'] = df_return['人気'].str.replace('人気','')
            df_return['払戻'] = df_return['払戻'].str.replace('円','').str.replace(',','').astype(int)
        
            #払い戻しを分ける
            if '前回騎乗' in kekka_df.columns:
                df_1 = df_return[df_return['券種'].isin(['単勝','複勝', '枠連'])].rename(columns={'券種':'馬番', '馬番':'馬名', '払戻':'馬齢', '人気':'オッズ'}).reset_index(drop=True)
                df_umaren = df_return[df_return['券種'].isin(['馬連'])].rename(columns={'券種':'馬番', '馬番':'馬名', '払戻':'馬齢', '人気':'オッズ'}).reset_index(drop=True)
                empty_row = pd.DataFrame([None]*df_1.shape[1]).T
                empty_row.columns = df_1.columns
                if '枠連' not in df_1['馬番'].values:
                    if int((df_1['馬番'] == '複勝').sum()) == 1:
                        for a in range(3):
                            df_1 = pd.concat([df_1, empty_row], ignore_index=True)
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'前回騎乗', '人気':'調教師'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                    elif int((df_1['馬番'] == '複勝').sum()) == 2:
                        for a in range(2):
                            df_1 = pd.concat([df_1, empty_row], ignore_index=True)
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'前回騎乗', '人気':'調教師'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                    elif int((df_1['馬番'] == '複勝').sum()) == 3:
                        df_1 = pd.concat([df_1, empty_row], ignore_index=True)
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'前回騎乗', '人気':'調教師'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                    else:
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'前回騎乗', '人気':'調教師'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                else:
                    df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                    df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'前回騎乗', '人気':'調教師'}).reset_index(drop=True)
                    df_3 = pd.concat([df_1,df_2], axis=1)
            else:
                df_1 = df_return[df_return['券種'].isin(['単勝','複勝', '枠連'])].rename(columns={'券種':'馬番', '馬番':'馬名', '払戻':'馬齢', '人気':'オッズ'}).reset_index(drop=True)
                df_umaren = df_return[df_return['券種'].isin(['馬連'])].rename(columns={'券種':'馬番', '馬番':'馬名', '払戻':'馬齢', '人気':'オッズ'}).reset_index(drop=True)
                empty_row = pd.DataFrame([None]*df_1.shape[1]).T
                empty_row.columns = df_1.columns
                if '枠連' not in df_1['馬番'].values:
                    if int((df_1['馬番'] == '複勝').sum()) == 1:
                        for a in range(3):
                            df_1 = pd.concat([df_1, empty_row], ignore_index=True)
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'調教師', '人気':'脚質'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                    elif int((df_1['馬番'] == '複勝').sum()) == 2:
                        for a in range(2):
                            df_1 = pd.concat([df_1, empty_row], ignore_index=True)
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'調教師', '人気':'脚質'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                    elif int((df_1['馬番'] == '複勝').sum()) == 3:
                        df_1 = pd.concat([df_1, empty_row], ignore_index=True)
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'調教師', '人気':'脚質'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                    else:
                        df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                        df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'調教師', '人気':'脚質'}).reset_index(drop=True)
                        df_3 = pd.concat([df_1,df_2], axis=1)
                else:
                    df_1 = pd.concat([df_1, df_umaren], ignore_index=True)
                    df_2 = df_return[df_return['券種'].isin(['ワイド','馬単','3連複', '3連単'])].rename(columns={'券種':'騎手名', '馬番':'斤量', '払戻':'調教師', '人気':'脚質'}).reset_index(drop=True)
                    df_3 = pd.concat([df_1,df_2], axis=1)
        except:
            data = {
            '馬番': ['単勝', '複勝', '複勝', '複勝', '枠連', '馬連'],
            '馬名': [0, 0, 0, 0, 0, 0],
            '馬齢': [0, 0, 0, 0, 0, 0],
            'オッズ': [0, 0, 0, 0, 0, 0],
            '騎手名': ['ワイド', 'ワイド', 'ワイド', '馬単', '3連複', '3連単'],
            '斤量': [0, 0, 0, 0, 0, 0],
            '調教師': [0, 0, 0, 0, 0, 0],
            '脚質': [0, 0, 0, 0, 0, 0]
            }
            df_3 = pd.DataFrame(data)
        #払い戻しと出馬、結果をconcatする
        df_sort = pd.concat([df_sort, df_3])
        kekka_df = pd.concat([kekka_df, df_3])
        shutuba_data_list.append(df_sort)
        kekka_df_list.append(kekka_df)
        url_list.append(url)
    #エクセルに保存
    print('エクセルに保存します')
    no_data_list = []
    book = openpyxl.Workbook()
    for date in tqdm(date_list):
        for kaijou in kaijou_data_list:
            for kaijou_date in kaijou_date_list:
                if len(str(date)) == 4:
                    if kaijou_date == kaijou+'_0'+str(date):
                        url_list_2 = []
                        for url, df_sort, kekka_df, shetname in zip(url_list, shutuba_data_list, kekka_df_list, url_list):
                            if kaijou == str(url.split('/')[4][:2]) and date == int(url.split('/')[4][-4:]):
                                shetname_list = ['出馬表'+'_'+ str(int(url.split('/')[4][9:11])),
                                        '結果_'+ str(int(url.split('/')[4][9:11]))]
                                #df_sort.to_excel(writer, sheet_name=shetname_list[0], index=False)
                                kekka_df.to_excel(shetname_list[1]+'.xlsx', sheet_name=shetname_list[1], engine='openpyxl', index=False)
                                url_list_2.append(url)
                        excel_color(date, kaijou, url_list_2)
                elif len(str(date)) == 3:
                    if kaijou_date == kaijou+'_00'+str(date):
                        url_list_2 = []
                        for url, df_sort, kekka_df, shetname in zip(url_list, shutuba_data_list, kekka_df_list, url_list):
                            if kaijou == str(url.split('/')[4][:2]) and date == int(url.split('/')[4][-4:]):
                                shetname_list = ['出馬表'+'_'+ str(int(url.split('/')[4][9:11])),
                                        '結果_'+ str(int(url.split('/')[4][9:11]))]
                                #df_sort.to_excel(writer, sheet_name=shetname_list[0], index=False)
                                kekka_df.to_excel(shetname_list[1]+'.xlsx', sheet_name=shetname_list[1], engine='openpyxl', index=False)
                                url_list_2.append(url)
                        excel_color(date, kaijou, url_list_2)
                else:
                    if kaijou_date == kaijou+'_'+str(date):
                        url_list_2 = []
                        for url, df_sort, kekka_df in zip(url_list, shutuba_data_list, kekka_df_list):
                            if kaijou == str(url.split('/')[4][:2]) and date == int(url.split('/')[4][-5:]):
                                shetname_list = ['出馬表'+'_'+ str(int(url.split('/')[4][9:11])),
                                        '結果_'+ str(int(url.split('/')[4][9:11]))]
                                #df_sort.to_excel(writer, sheet_name=shetname_list[0], index=False)
                                kekka_df.to_excel(shetname_list[1]+'.xlsx', sheet_name=shetname_list[1], engine='openpyxl', index=False)
                                url_list_2.append(url)
                        excel_color(date, kaijou, url_list_2)

# %%
for month in tqdm(range(1,13)):
    print(str(year)+str(month).zfill(2)+'月の出馬表を取得します')
    make_excel_file(str(year)+str(month).zfill(2))

# %%
